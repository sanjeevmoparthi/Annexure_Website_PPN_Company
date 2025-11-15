from io import BytesIO
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import io

# =====================================================
#                   Annexure – 1
# =====================================================
def annexure1_generate_excel_bytes(df: pd.DataFrame) -> BytesIO:

    required_cols = [
        'Branch', 'Vendor Name', 'Product Department',
        'MRP', 'Sold Qty', 'Sold Value', 'Total LandedCost'
    ]

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    df['Profit'] = df['Sold Value'] - df['Total LandedCost']

    summary = (
        df.groupby(['Branch', 'Vendor Name', 'Product Department'], as_index=False)
        .agg({
            'MRP': 'sum',
            'Sold Qty': 'sum',
            'Sold Value': 'sum',
            'Total LandedCost': 'sum',
            'Profit': 'sum'
        })
    )

    summary['Margin (%)'] = ((summary['Profit'] / summary['Sold Value']) * 100).round(2)

    branch_sales = summary.groupby('Branch')['Sold Value'].sum().sort_values(ascending=False)
    summary['Branch'] = pd.Categorical(
        summary['Branch'],
        categories=branch_sales.index.tolist(),
        ordered=True
    )
    summary.sort_values(by=['Branch', 'Sold Value'], ascending=[True, False], inplace=True)

    wb = Workbook()
    std = wb.active
    wb.remove(std)

    out = BytesIO()

    for dept, dept_df in summary.groupby("Product Department"):
        ws = wb.create_sheet(title=str(dept)[:31])

        headers = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
            f"DEPARTMENT - {str(dept).upper()}",
            "Annexure - I",
            "Vendor Wise Margin",
            "(Amount in Rs.)"
        ]

        for row_idx, text in enumerate(headers, start=1):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
            cell = ws.cell(row=row_idx, column=1, value=text)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True, size=12)

        start_row = len(headers) + 2

        table_cols = [
            'Branch', 'Vendor Name', 'Product Department',
            'MRP', 'Sold Qty', 'Sold Value',
            'Total LandedCost', 'Profit', 'Margin (%)'
        ]

        for col_idx, col_name in enumerate(table_cols, start=1):
            ws.cell(row=start_row, column=col_idx, value=col_name).font = Font(bold=True)

        for r_idx, row in enumerate(dataframe_to_rows(dept_df[table_cols], index=False, header=False),
                                    start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        for i, col in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 3

    wb.save(out)
    out.seek(0)
    return out


# =====================================================
#                   Annexure – 2
# =====================================================
def annexure2_generate_excel_bytes(df):

    df.columns = df.columns.str.strip()

    required_cols = [
        "Branch", "Brand", "Product Department",
        "MRP", "Sold Qty", "Sold Value", "Total LandedCost", "Profit"
    ]

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise Exception(f"Missing required columns: {missing}")

    if "Profit" not in df.columns:
        df["Profit"] = df["Sold Value"] - df["Total LandedCost"]

    df["Margin %"] = (df["Profit"] / df["Sold Value"]) * 100
    df["Margin %"] = df["Margin %"].round(2)

    summary = (
        df.groupby(["Branch", "Brand", "Product Department"], as_index=False)
        .agg({
            "MRP": "sum",
            "Sold Qty": "sum",
            "Sold Value": "sum",
            "Total LandedCost": "sum",
            "Profit": "sum",
            "Margin %": "mean"
        })
    )

    summary = summary.sort_values(by=["Branch", "Sold Value"], ascending=[True, False])

    wb = Workbook()
    wb.remove(wb.active)

    for dept, data in summary.groupby("Product Department"):

        ws = wb.create_sheet(title=str(dept)[:31])

        headers = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 to 31-OCT-2025",
            f"DEPARTMENT - {dept}",
            "Annexure-II",
            "Brand Wise Margin",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(headers, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=9)
            cell = ws.cell(row=i, column=1, value=text)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center")

        for row in dataframe_to_rows(data, index=False, header=True):
            ws.append(row)

        for i, col in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 2

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# =====================================================
#                   Annexure – 3
# =====================================================
def annexure3_generate_excel_bytes(df):

    df.columns = df.columns.str.strip()

    required_cols = ["Branch", "Brand", "Product Department", "Sold Qty", "Sold Value"]

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    summary = (
        df.groupby(["Branch", "Brand", "Product Department"], as_index=False)
        .agg({"Sold Qty": "sum", "Sold Value": "sum"})
    )

    wb = Workbook()
    wb.remove(wb.active)

    for dept, data in summary.groupby("Product Department"):

        branch_order = (
            data.groupby("Branch")["Sold Value"]
            .sum()
            .sort_values(ascending=False)
            .index
            .tolist()
        )

        data["Branch"] = pd.Categorical(data["Branch"], categories=branch_order, ordered=True)
        data = data.sort_values(["Branch", "Sold Value"], ascending=[True, False]).reset_index(drop=True)

        data["Rank"] = (
            data.groupby("Branch")["Sold Value"]
            .rank(method="dense", ascending=False)
            .astype(int)
        )

        ws = wb.create_sheet(title=str(dept)[:31])

        headers = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 to 31-OCT-2025",
            f"DEPARTMENT - {dept}",
            "Annexure - III",
            "Brand Wise Sales Quantity & Value",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(headers, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
            cell = ws.cell(row=i, column=1, value=text)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center")

        for row in dataframe_to_rows(data, index=False, header=True):
            ws.append(row)

        for i, col in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value)) for c in col if c.value not in (None, "")), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 3

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# =====================================================
#                   Annexure – 4
# =====================================================
def annexure4_generate_excel_bytes(df):

    df.columns = df.columns.str.strip().str.lower()

    required_cols = [
        "branch", "brand", "product category",
        "product department", "sold qty", "sold value", "profit"
    ]

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise KeyError(f"Missing columns in Excel: {missing}")

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine="openpyxl")

    for dept, df_dept in df.groupby("product department"):

        summary = (
            df_dept.groupby(["branch", "brand", "product category"], as_index=False)
            .agg({
                "sold qty": "sum",
                "sold value": "sum",
                "profit": "sum"
            })
        )

        branch_totals = summary.groupby("branch", as_index=False)["sold value"] \
            .sum().rename(columns={"sold value": "branch_total"})

        brand_totals = summary.groupby(["branch", "brand"], as_index=False)["sold value"] \
            .sum().rename(columns={"sold value": "brand_total"})

        summary = summary.merge(branch_totals, on="branch", how="left")
        summary = summary.merge(brand_totals, on=["branch", "brand"], how="left")

        summary["Rank"] = (
            summary.groupby(["branch", "brand"])["sold value"]
            .rank(method="dense", ascending=False)
            .astype(int)
        )

        summary = summary.sort_values(
            by=["branch_total", "branch", "brand_total", "brand", "sold value"],
            ascending=[False, True, False, True, False]
        ).reset_index(drop=True)

        summary = summary[[
            "branch", "brand", "product category",
            "sold qty", "sold value", "Rank"
        ]]

        safe_sheet = str(dept).replace("/", "_")[:31]
        summary.to_excel(writer, index=False, sheet_name=safe_sheet, startrow=7)

        ws = writer.book[safe_sheet]

        header_texts = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
            f"DEPARTMENT - {dept.upper()}",
            "Annexure - IV",
            "Product Wise Sales Quantity And Value",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(header_texts, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
            cell = ws.cell(row=i, column=1)
            cell.value = text
            cell.font = Font(name="Calibri", size=12, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    writer.close()
    output.seek(0)
    return output

# =====================================================
#                   Annexure – 5
# =====================================================
def annexure5_generate_excel_bytes(df):

    import pandas as pd
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    # ----------------------------
    # Step 1: Clean columns
    # ----------------------------
    df.columns = df.columns.str.strip().str.lower()

    # ----------------------------
    # Step 2: Validate required columns
    # ----------------------------
    required_cols = [
        "branch", "product category", "product department",
        "sold qty", "sold value"
    ]

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise KeyError(f"Missing columns: {missing}")

    # ----------------------------
    # Step 3: Create Workbook
    # ----------------------------
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ----------------------------
    # Step 4: Loop Department-wise
    # ----------------------------
    for dept, df_dept in df.groupby("product department"):

        # --- 4.1: Aggregate ---
        summary = (
            df_dept.groupby(["branch", "product category"], as_index=False)
            .agg({
                "sold qty": "sum",
                "sold value": "sum"
            })
        )

        # --- 4.2: Branch-wise total ---
        branch_total = (
            summary.groupby("branch", as_index=False)["sold value"]
            .sum()
            .rename(columns={"sold value": "branch_total"})
        )

        summary = summary.merge(branch_total, on="branch", how="left")

        # --- 4.3: % Contribution ---
        summary["%Contribution"] = (
            summary["sold value"] / summary["branch_total"] * 100
        ).map(lambda x: f"{x:.2f}%")

        # --- 4.4: Sorting ---
        summary = summary.sort_values(
            by=["branch_total", "branch", "sold value"],
            ascending=[False, True, False]
        ).reset_index(drop=True)

        # --- 4.5: Final Column Order ---
        summary = summary[
            ["branch", "product category", "sold qty", "sold value", "%Contribution"]
        ]

        # ----------------------------
        # Create Sheet
        # ----------------------------
        sheet_name = str(dept).replace("/", "_").replace("\\", "_")[:31]
        ws = wb.create_sheet(title=sheet_name)

        # ----------------------------
        # Header Section
        # ----------------------------
        header_texts = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
            f"DEPARTMENT - {dept.upper()}",
            "Annexure - V",
            "Product Category Contribution - All Branches",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(header_texts, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
            cell = ws.cell(row=i, column=1)
            cell.value = text
            cell.font = Font(name="Calibri", size=12, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ----------------------------
        # Data Table Start Row
        # ----------------------------
        start_row = 8

        # --- Write header row ---
        for col_idx, col_name in enumerate(summary.columns, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.font = Font(bold=True)

        # --- Write data rows ---
        for row_idx, row in enumerate(summary.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # --- Auto column width ---
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3

    # ----------------------------
    # RETURN EXCEL BYTES
    # ----------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
# =====================================================
#                   Annexure – 6
# =====================================================
# import pandas as pd
# import numpy as np
# from io import BytesIO
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Font, Alignment

def annexure6_generate_excel_bytes(df):
    """
    Annexure VI - Negative Margin Report
    Returns a multi-sheet Excel file (one sheet per department) as BytesIO.
    """
    
    # ---------- Normalize Columns ----------
    df.columns = df.columns.str.strip().str.lower()

    def find_col(possible):
        for c in df.columns:
            if c in possible:
                return c
        return None

    item_code_col = find_col(["item code", "itemcode"])
    product_name_col = find_col(["product name", "productname"])
    branch_col = find_col(["branch"])
    mrp_col = find_col(["mrp"])
    sold_qty_col = find_col(["sold qty", "soldqty"])
    sold_value_col = find_col(["sold value", "soldvalue"])
    landed_col = find_col(["total landedcost", "total landed cost"])
    dept_col = find_col(["product department", "productdepartment"])

    if not all([sold_qty_col, sold_value_col, landed_col, branch_col]):
        raise KeyError("Missing columns: Sold Qty, Sold Value, Landed Cost, Branch")

    if dept_col is None:
        df["product department"] = "All"
        dept_col = "product department"

    # ---------- Convert numeric ----------
    for col in [sold_qty_col, sold_value_col, landed_col, mrp_col]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # ---------- Compute Metrics ----------
    df["Sales per Qty"] = np.where(df[sold_qty_col] != 0,
                                   df[sold_value_col] / df[sold_qty_col], 0)
    df["L Cost per Qty"] = np.where(df[sold_qty_col] != 0,
                                    df[landed_col] / df[sold_qty_col], 0)
    df["Total Loss"] = df[sold_value_col] - df[landed_col]

    # ---------- Filter Negative Margin ----------
    neg = df[df["L Cost per Qty"] > df["Sales per Qty"]].copy()

    # Excel Output
    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    if neg.empty:
        ws = wb.create_sheet("No Negative Margin")
        ws["A1"] = "No Negative Margin Records Found"
        ws["A1"].font = Font(bold=True)
    else:
        # Create sheet per department
        for dept, g in neg.groupby(dept_col):
            ws = wb.create_sheet(str(dept)[:31])

            g = g.reset_index(drop=True)
            g["S.No"] = range(1, len(g) + 1)

            final = pd.DataFrame({
                "S.No": g["S.No"],
                "Item Code": g.get(item_code_col, ""),
                "Product Name": g.get(product_name_col, ""),
                "Branch": g[branch_col],
                "MRP": g[mrp_col],
                "Sales per Qty": g["Sales per Qty"].round(2),
                "L Cost per Qty": g["L Cost per Qty"].round(2),
                "Sold Qty": g[sold_qty_col],
                "Sold Value": g[sold_value_col].round(2),
                "Total Landed Cost": g[landed_col].round(2),
                "Total Loss": g["Total Loss"].round(2)
            })

            # --- Headings ---
            headers = [
                "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
                "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
                f"DEPARTMENT - {dept.upper()}",
                "Annexure - VI",
                "Selling Price Less Than Purchase Cost – Negative Margin Sales",
                "(Amount in Rs.)"
            ]

            for i, text in enumerate(headers, start=1):
                ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(final.columns))
                c = ws.cell(row=i, column=1, value=text)
                c.font = Font(bold=True, size=12)
                c.alignment = Alignment(horizontal="center")

            # Column Headers
            start_row = len(headers) + 2
            for j, col in enumerate(final.columns, start=1):
                cell = ws.cell(row=start_row, column=j, value=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Data Rows
            for r_idx, row in enumerate(final.itertuples(index=False), start=start_row + 1):
                for c_idx, val in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

            # Auto column width
            for i, col_cells in enumerate(ws.columns, start=1):
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
                ws.column_dimensions[get_column_letter(i)].width = max_len + 3

    # Save to BytesIO
    wb.save(output)
    output.seek(0)
    return output


# Annexure_function.py with Annexure 7 added

# Paste your existing Annexure 1–6 code above this section

# =====================================================
#                  Annexure – 7
# =====================================================
def annexure7_generate_excel_bytes(df):
    import pandas as pd
    import numpy as np
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    # Normalize column names
    df = df.copy()
    df.columns = df.columns.str.strip().str.lower()

    # Helper to find column
    def find_col(possible):
        for c in df.columns:
            if c in possible:
                return c
        return None

    item_code_col = find_col(["item code", "itemcode"])
    product_name_col = find_col(["product name", "productname"])
    branch_col = find_col(["branch"])
    mrp_col = find_col(["mrp"])
    sold_qty_col = find_col(["sold qty", "soldqty"])
    sold_value_col = find_col(["sold value", "soldvalue"])
    landed_col = find_col(["total landedcost", "total landed cost"])
    profit_col = find_col(["profit"])
    dept_col = find_col(["product department", "productdepartment"])

    if not all([sold_qty_col, sold_value_col, landed_col, profit_col, branch_col]):
        raise KeyError("Missing one or more required columns: Sold Qty, Sold Value, Landed Cost, Profit, Branch")

    if dept_col is None:
        df["product department"] = "All"
        dept_col = "product department"

    # Convert numeric fields
    for col in [sold_qty_col, sold_value_col, landed_col, profit_col, mrp_col]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Derived metrics
    df["Sales per Qty"] = np.where(df[sold_qty_col] != 0, df[sold_value_col] / df[sold_qty_col], 0)
    df["L Cost per Qty"] = np.where(df[sold_qty_col] != 0, df[landed_col] / df[sold_qty_col], 0)
    df["Profit % on Sales"] = np.where(df[sold_value_col] != 0, (df[profit_col] / df[sold_value_col]) * 100, 0)

    # Filter Profit > 0 and Profit% < 10
    filtered = df[(df[profit_col] > 0) & (df["Profit % on Sales"] < 10)].copy()

    if filtered.empty:
        raise ValueError("No Products Found with Profit Below 10% and Greater Than 0.")

    # Prepare output workbook (multi-sheet)
    wb = Workbook()
    wb.remove(wb.active)

    for dept, g in filtered.groupby(dept_col):
        g = g.reset_index(drop=True)
        g["S.No"] = range(1, len(g) + 1)

        final = pd.DataFrame({
            "S.No": g["S.No"],
            "Item Code": g.get(item_code_col, ""),
            "Product Name": g.get(product_name_col, ""),
            "Branch": g[branch_col],
            "MRP": g[mrp_col].round(2),
            "Sales per Qty": g["Sales per Qty"].round(2),
            "L Cost per Qty": g["L Cost per Qty"].round(2),
            "Sold Qty": g[sold_qty_col].round(2),
            "Sold Value": g[sold_value_col].round(2),
            "Total Landed Cost": g[landed_col].round(2),
            "Profit": g[profit_col].round(2),
            "Profit % on Sales": g["Profit % on Sales"].round(2)
        })

        # Sheet creation
        safe_name = str(dept).replace("/", "_").replace("\\", "_")[:31]
        ws = wb.create_sheet(title=safe_name)

        headers = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
            f"DEPARTMENT - {dept.upper()}",
            "Annexure - VII",
            "List of Products Sold at a Profit Below 10%",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(headers, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(final.columns))
            c = ws.cell(row=i, column=1, value=text)
            c.font = Font(bold=True, size=12)
            c.alignment = Alignment(horizontal="center", vertical="center")

        start_row = len(headers) + 2

        # Column headers
        for j, col in enumerate(final.columns, start=1):
            cell = ws.cell(row=start_row, column=j, value=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for r_idx, row in enumerate(final.itertuples(index=False), start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Auto column width
        for i, col_cells in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(i)
            max_len = max(len(str(cell.value)) for cell in col_cells if cell.value)
            ws.column_dimensions[col_letter].width = max_len + 3

    # Return file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def annexure8_generate_excel_bytes(df):
    """
    Annexure VIII – Products sold with Neither Profit Nor Loss
    Returns an Excel file (BytesIO) containing multiple sheets:
    One sheet per department.
    """

    import pandas as pd
    import numpy as np
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    # Work on a copy
    df = df.copy()
    df.columns = df.columns.str.strip().str.lower()

    # ---- helper to find column ----
    def find_col(possible):
        for c in df.columns:
            if c in possible:
                return c
        return None

    # Identify columns
    item_code_col = find_col(["item code", "itemcode"])
    product_name_col = find_col(["product name", "productname"])
    branch_col = find_col(["branch"])
    mrp_col = find_col(["mrp"])
    sold_qty_col = find_col(["sold qty", "soldqty"])
    sold_value_col = find_col(["sold value", "soldvalue"])
    landed_col = find_col(["total landedcost", "total landed cost"])
    dept_col = find_col(["product department", "productdepartment"])

    if not all([sold_qty_col, sold_value_col, landed_col, branch_col]):
        raise KeyError(
            "Missing one or more required columns: Sold Qty, Sold Value, Landed Cost, Branch"
        )

    if dept_col is None:
        df["product department"] = "All"
        dept_col = "product department"

    # numeric conversion
    for col in [sold_qty_col, sold_value_col, landed_col, mrp_col]:
        if col:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # derived metrics
    df["sales_per_qty"] = np.where(df[sold_qty_col] != 0,
                                   df[sold_value_col] / df[sold_qty_col],
                                   0)
    df["lcost_per_qty"] = np.where(df[sold_qty_col] != 0,
                                   df[landed_col] / df[sold_qty_col],
                                   0)

    # Filter: Neither Profit Nor Loss
    filtered = df[
        (df[sold_qty_col] > 0) &
        (np.isclose(df[sold_value_col], df[landed_col], atol=0.01))
    ].copy()

    # create workbook
    wb = Workbook()
    wb.remove(wb.active)  # clean default sheet

    if filtered.empty:
        # create one sheet saying no data
        ws = wb.create_sheet("No Data")
        ws["A1"] = "No Products Found with Neither Profit Nor Loss"
        ws["A1"].font = Font(bold=True)
    else:
        # one sheet per department
        for dept, g in filtered.groupby(dept_col):
            g = g.reset_index(drop=True)
            g["S.No"] = range(1, len(g) + 1)

            final = pd.DataFrame({
                "S.No": g["S.No"],
                "Item Code": g.get(item_code_col, ""),
                "Product Name": g.get(product_name_col, ""),
                "Branch": g[branch_col],
                "MRP": g[mrp_col].astype(int),
                "Sales per Qty": g["sales_per_qty"].astype(int),
                "L Cost per Qty": g["lcost_per_qty"].astype(int),
                "Sold Qty": g[sold_qty_col].astype(int),
                "Sold Value": g[sold_value_col].astype(int),
                "Total Landed Cost": g[landed_col].astype(int)
            })

            ws = wb.create_sheet(title=str(dept)[:31])

            # headers
            headers = [
                "POTHYS RETAIL PRIVATE LIMITED - All BRANCH",
                "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
                f"DEPARTMENT - {str(dept).upper()}",
                "Annexure - VIII",
                "Products sold with neither profit nor loss",
                "(Amount in Rs.)"
            ]

            for i, text in enumerate(headers, start=1):
                ws.merge_cells(start_row=i, start_column=1,
                               end_row=i, end_column=len(final.columns))
                c = ws.cell(row=i, column=1, value=text)
                c.font = Font(bold=True, size=12)
                c.alignment = Alignment(horizontal="center", vertical="center")

            # column headers
            start_row = len(headers) + 2
            for j, colname in enumerate(final.columns, start=1):
                cell = ws.cell(row=start_row, column=j, value=colname)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # data
            for r_idx, row in enumerate(final.itertuples(index=False), start=start_row + 1):
                for c_idx, val in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

            # auto column width
            for col_idx in range(1, final.shape[1] + 1):
                column_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws[column_letter]:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = max_len + 3

    # output as bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# =====================================================
#                   Annexure – 9
# =====================================================
def annexure9_generate_excel_bytes(df):
    """
    Annexure IX – High Vendor Margin but Less Profit Margin
    Returns an Excel file (BytesIO) containing multiple sheets: One per department
    """
    import pandas as pd
    import numpy as np
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    df = df.copy()
    df.columns = df.columns.str.strip().str.lower()

    # Helper to find column
    def find_col(possible):
        for c in df.columns:
            if c in possible:
                return c
        return None

    # Identify columns
    item_code_col = find_col(["item code", "itemcode"])
    product_name_col = find_col(["product name", "productname"])
    branch_col = find_col(["branch"])
    mrp_col = find_col(["mrp"])
    sold_qty_col = find_col(["sold qty", "soldqty"])
    sold_value_col = find_col(["sold value", "soldvalue"])
    landed_col = find_col(["total landedcost", "total landed cost"])
    profit_col = find_col(["profit"])
    dept_col = find_col(["product department", "productdepartment"])

    if not all([sold_qty_col, sold_value_col, landed_col, profit_col, branch_col]):
        raise KeyError("Missing one or more required columns: Sold Qty, Sold Value, Landed Cost, Profit, Branch")

    if dept_col is None:
        df["product department"] = "All"
        dept_col = "product department"

    # Numeric conversion
    for col in [sold_qty_col, sold_value_col, landed_col, profit_col, mrp_col]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Derived metrics
    df["Profit per Unit"] = np.where(df[sold_qty_col] != 0, df[profit_col] / df[sold_qty_col], 0)
    df["Sales per Qty"] = np.where(df[sold_qty_col] != 0, df[sold_value_col] / df[sold_qty_col], 0)
    df["L Cost per Qty"] = np.where(df[sold_qty_col] != 0, df[landed_col] / df[sold_qty_col], 0)
    df["Vendor Margin %"] = np.where(df[mrp_col] != 0, (df[mrp_col] - df["L Cost per Qty"]) / df[mrp_col] * 100, 0)
    df["Profit Margin %"] = np.where(df["Sales per Qty"] != 0, df["Profit per Unit"] / df["Sales per Qty"] * 100, 0)

    # Filter: Vendor Margin > Profit Margin and difference > 1%
    filtered = df[(df["Vendor Margin %"] - df["Profit Margin %"]) > 1].copy()
    if filtered.empty:
        raise ValueError("No records found for Annexure IX criteria.")

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    for dept, g in filtered.groupby(dept_col):
        g = g.reset_index(drop=True)
        g["S.No"] = range(1, len(g) + 1)

        final = pd.DataFrame({
            "S.No": g["S.No"],
            "Item Code": g.get(item_code_col, ""),
            "Product Name": g.get(product_name_col, ""),
            "Branch": g[branch_col],
            "MRP": g[mrp_col].round(2),
            "Sales per Qty": g["Sales per Qty"].round(2),
            "L Cost per Qty": g["L Cost per Qty"].round(2),
            "Profit per Unit": g["Profit per Unit"].round(2),
            "Sold Qty": g[sold_qty_col].round(2),
            "Vendor Margin %": g["Vendor Margin %"].round(2),
            "Profit Margin %": g["Profit Margin %"].round(2)
        })

        ws = wb.create_sheet(title=str(dept)[:31])

        headers = [
            "POTHYS RETAIL PRIVATE LIMITED - ALL BRANCH",
            "INTERNAL AUDIT FOR THE PERIOD 01-OCT-2025 TO 31-OCT-2025",
            f"DEPARTMENT - {dept.upper()}",
            "Annexure - IX",
            "High Vendor Margin but Less Profit Margin",
            "(Amount in Rs.)"
        ]

        for i, text in enumerate(headers, start=1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(final.columns))
            c = ws.cell(row=i, column=1, value=text)
            c.font = Font(bold=True, size=12)
            c.alignment = Alignment(horizontal="center", vertical="center")

        start_row = len(headers) + 2
        for j, colname in enumerate(final.columns, start=1):
            cell = ws.cell(row=start_row, column=j, value=colname)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for r_idx, row in enumerate(final.itertuples(index=False), start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Auto column width
        for i, col_cells in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(i)
            max_len = max(len(str(cell.value)) for cell in col_cells if cell.value)
            ws.column_dimensions[col_letter].width = max_len + 3

    # Return Excel bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

